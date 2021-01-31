import argparse
import csv
import io
import requests
import sys

from bs4 import BeautifulSoup
from openpyxl import load_workbook

DEFAULT_URL = 'https://www.mass.gov/info-details/covid-19-vaccine-locations-for-individuals-in-phase-1#find-a-location-to-get-vaccinated-if-eligible-'

# Class attribute of relevant links to search
MA_CALLOUT_LINK_CLASS = 'ma__callout-link'


parser = argparse.ArgumentParser(description='Pull Vaccine Site Info from MA')
parser.add_argument(
    '--url',
    nargs=1,
    default=DEFAULT_URL,
    help=f'Site to scrape. Defaults to {DEFAULT_URL}'
)
parser.add_argument(
    '--outfile',
    '-o',
    nargs='?',
    type=argparse.FileType('w'),
    default=sys.stdout,
    help='Filename to write downloaded csv to. Defaults to stdout'
)
parser.add_argument(
    '--link-match-checks',
    '-l',
    nargs='+',
    action='extend',
    default=['download'],
    help='Strings to match against the title of the link to search for. Defaults to ["download"] (Capitalization does not matter)'
)

args = parser.parse_args()
response = requests.get(args.url)
# Parse the response from MA site
soup = BeautifulSoup(response.text, "html.parser")
# Find all links with the appropriate class
relevant_links = soup.find_all('a', {'class': MA_CALLOUT_LINK_CLASS})

text_to_url = {}
link_texts = []
text_match_checks = [match_check.lower() for match_check in args.link_match_checks]
for link in relevant_links:
    link_text = str.lower('\n'.join(link.stripped_strings))
    link_texts.append(link_text)
    # Check if text displayed with the link matches the input texts to match
    # it against
    matches = all([match_check in link_text for match_check in text_match_checks])
    if not matches:
        continue
    text_to_url[link_text] = link['href']

assert len(text_to_url) > 0, f'No links found matching {text_match_checks}. All link texts: {link_texts}'
assert len(text_to_url) < 2, f'Multiple links match {text_match_checks}: {text_to_url}'

# Only one url in text_to_url, but dict_values aren't subscritable
vaccination_site_url = [u for u in text_to_url.values()][0]
# Grab the xlsx file from the url we scraped
vaccination_site_resp = requests.get(vaccination_site_url)
# Load the xlsx file into a stream and use openpyxl to open it
# Will return an error if the file is not xlsx
tmp_xlsx = io.BytesIO(vaccination_site_resp.content)
work_book = load_workbook(tmp_xlsx)
# Use the active sheet of the excel workbook - there should only be one
work_sheet = work_book.active
outfile = csv.writer(args.outfile)
for row in work_sheet.iter_rows():
    to_write = [cell.value for cell in row]
    # Skip first line if it's a large header to make csv more parsable
    if to_write == ['Vaccine locations']:
        continue
    outfile.writerow(to_write)
