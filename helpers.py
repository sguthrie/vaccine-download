import argparse
import io
import os
import requests
import time
import sys

from airtable import Airtable
from bs4 import BeautifulSoup
from collections import defaultdict
from openpyxl import load_workbook

# Class attribute of relevant links to search
MA_CALLOUT_LINK_CLASS = 'ma__callout-link'

def get_download_link(ma_url, link_match_checks):
    response = requests.get(ma_url)
    # Parse the response from MA site
    soup = BeautifulSoup(response.text, "html.parser")
    # Find all links with the appropriate class
    relevant_links = soup.find_all('a', {'class': MA_CALLOUT_LINK_CLASS})

    text_to_url = {}
    link_texts = []
    text_match_checks = [match_check.lower() for match_check in link_match_checks]
    for link in relevant_links:
        link_text = str.lower('\n'.join(link.stripped_strings))
        link_texts.append(link_text)
        # Check if text displayed with the link matches the input texts to match
        # it against
        matches = all([match_check in link_text for match_check in text_match_checks])
        if not matches:
            continue
        text_to_url[link_text] = link['href']

    assert len(text_to_url) > 0, f'No links found matching {text_match_checks}. All link texts: {link_texts}'
    assert len(text_to_url) < 2, f'Multiple links match {text_match_checks}: {text_to_url}'

    # Only one url in text_to_url, but dict_values aren't subscritable
    return [u for u in text_to_url.values()][0]

def get_curr_ma_data(vaccination_site_url, airtable_unique_val):
    # Grab the xlsx file from the url we scraped
    vaccination_site_resp = requests.get(vaccination_site_url)
    # Load the xlsx file into a stream and use openpyxl to open it
    # Will return an error if the file is not xlsx
    tmp_xlsx = io.BytesIO(vaccination_site_resp.content)
    work_book = load_workbook(tmp_xlsx)
    # Use the active sheet of the excel workbook - there should only be one
    work_sheet = work_book.active
    curr_ma_data = {}
    skipped_first = False
    for i, row in enumerate(work_sheet.iter_rows()):
        to_write = [cell.value.strip() if cell.value else None for cell in row]
        # Skip first line if it's a large header to make csv more parsable
        if i == 0 and to_write[0] == 'Vaccine locations':
            skipped_first = True
            continue
        elif i == 0:
            # Grab header columns for data
            header = to_write
            try:
                name_column = header.index(airtable_unique_val)
            except ValueError as e:
                print(header)
                raise e
            continue
        elif i == 1 and skipped_first:
            header = to_write
            try:
                name_column = header.index(airtable_unique_val)
            except ValueError as e:
                print(header)
                raise e
            continue
        if to_write[name_column] in curr_ma_data:
            raise RuntimeError('Duplicate names in data')
        curr_ma_data[to_write[name_column]] = {
            h: v for h, v in zip(header, to_write)
        }
    return curr_ma_data

def get_curr_airtable_data_and_update_list(table, airtable_unique_val, curr_ma_data):
    curr_airtable_data = {}
    to_update = defaultdict(list)
    table_results = table.get_all()
    for entry in table_results:
        entry_name = entry['fields'][airtable_unique_val].strip()
        if entry_name in curr_airtable_data:
            raise RuntimeError('Expect names to be unique')
        curr_airtable_data[entry_name] = entry
        if entry_name in curr_ma_data:
            for header in curr_ma_data[entry_name]:
                if header not in entry['fields']:
                    to_update[entry_name].append(header)
                elif entry['fields'][header] != curr_ma_data[entry_name][header]:
                    to_update[entry_name].append(header)
    return curr_airtable_data, to_update
